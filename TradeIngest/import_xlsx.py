"""
One-time importer: seed legacy_trade_id and trade_journal from
TraderJournal/trades.xlsx (the canonical journal file).

That file is a superset: TIDs 1038..3477 covering 2024-12 through 2026.
The only trades it doesn't have (3478..3482, the 5 newest scratches)
have no notes anyway -- they stay as empty journal rows in Postgres.

This importer:
  1. Reads TraderJournal/trades.xlsx (Executions + Trades tabs).
  2. Matches each xlsx row to a Postgres `trades` row on
     (date, symbol, entry_time).
  3. Re-stamps `trades.legacy_trade_id` to the xlsx Trade ID so
     historical IDs match what the user knows. Trades that ingested
     into Postgres but aren't in xlsx (e.g. swing-trade entries from
     2024-12 anchored before xlsx coverage) keep their auto-assigned
     placeholder IDs.
  4. Upserts `trade_journal` keyed by the (rewritten) legacy_trade_id.

Default mode is DRY-RUN. Pass --commit to write.

Usage:
    python import_xlsx.py              # dry-run: report only
    python import_xlsx.py --commit     # actually write
    python import_xlsx.py --verbose    # show every match
"""

import os
import sys
from collections import defaultdict
from datetime import date, time, datetime
from pathlib import Path

import openpyxl
import psycopg2
from dotenv import load_dotenv

SCRIPT_DIR = Path(__file__).parent
ENV_PATH = SCRIPT_DIR / ".env"
XLSX_PATH = SCRIPT_DIR.parent / "TraderJournal" / "trades.xlsx"
TARGET_DB = "trades_db"


# ---------------------------------------------------------------------------
# Connection
# ---------------------------------------------------------------------------

def get_conn():
    load_dotenv(ENV_PATH)
    return psycopg2.connect(
        host=os.environ["PG_HOST"],
        port=int(os.environ["PG_PORT"]),
        user=os.environ["PG_USER"],
        password=os.environ["PG_PASSWORD"],
        dbname=TARGET_DB,
    )


# ---------------------------------------------------------------------------
# xlsx readers
# ---------------------------------------------------------------------------

# Excel epoch: 1899-12-30. Day 1 = 1900-01-01. (Lotus 1-2-3 leap-year bug means
# this avoids the pre-March-1900 corner case.)
_EXCEL_EPOCH = date(1899, 12, 30)


def to_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        from datetime import timedelta
        return _EXCEL_EPOCH + timedelta(days=int(v))
    return None


def to_time(v):
    if isinstance(v, time):
        return v
    if isinstance(v, datetime):
        return v.time()
    if isinstance(v, (int, float)):
        # Excel serial time: fraction of a day. 0.5 == 12:00:00.
        total = round(float(v) * 86400)
        h, rem = divmod(total, 3600)
        m, s = divmod(rem, 60)
        return time(h, m, s)
    return None


def to_x_flag(v):
    """X-flags can be 0, 0.5, or 1 (xlsx sometimes assigns partial credit)."""
    if v is None or v == "":
        return 0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0


def to_float_or_none(v):
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (ValueError, TypeError):
        return None


def to_str_or_none(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def read_executions(ws):
    """One row per trade in the xlsx Executions tab. Return dict keyed by Trade ID."""
    out = {}
    for r in range(2, ws.max_row + 1):
        tid = ws.cell(row=r, column=1).value
        if tid is None:
            continue
        out[int(tid)] = {
            "trade_id":         int(tid),
            "date":             to_date(ws.cell(row=r, column=2).value),
            "entry_time":       to_time(ws.cell(row=r, column=3).value),
            "symbol":           to_str_or_none(ws.cell(row=r, column=4).value),
            "side":             to_str_or_none(ws.cell(row=r, column=5).value),
            "dollar_risk":      to_float_or_none(ws.cell(row=r, column=24).value),
            "x_failing_goal":   to_x_flag(ws.cell(row=r, column=27).value),
            "x_non_playbook":   to_x_flag(ws.cell(row=r, column=28).value),
            "x_selection":      to_x_flag(ws.cell(row=r, column=29).value),
            "x_entry":          to_x_flag(ws.cell(row=r, column=30).value),
            "x_sizing":         to_x_flag(ws.cell(row=r, column=31).value),
            "x_exit":           to_x_flag(ws.cell(row=r, column=32).value),
            "x_emotional":      to_x_flag(ws.cell(row=r, column=33).value),
            "x_preparation":    to_x_flag(ws.cell(row=r, column=34).value),
            "_row":             r,
        }
    return out


def read_trades(ws):
    """One row per trade in the xlsx Trades tab. Return dict keyed by Trade ID."""
    out = {}
    for r in range(2, ws.max_row + 1):
        tid = ws.cell(row=r, column=1).value
        if tid is None:
            continue
        out[int(tid)] = {
            "trade_id":      int(tid),
            "setup":         to_str_or_none(ws.cell(row=r, column=12).value),
            "sub_setup":     to_str_or_none(ws.cell(row=r, column=13).value),
            "trigger":       to_str_or_none(ws.cell(row=r, column=14).value),
            "tags":          to_str_or_none(ws.cell(row=r, column=15).value),
            "entry_notes":   to_str_or_none(ws.cell(row=r, column=16).value),
            "exit_notes":    to_str_or_none(ws.cell(row=r, column=17).value),
            "notes":         to_str_or_none(ws.cell(row=r, column=18).value),
            "mistake_notes": to_str_or_none(ws.cell(row=r, column=19).value),
            "chart":         to_str_or_none(ws.cell(row=r, column=20).value),
            "win_xlsx":      ws.cell(row=r, column=9).value,
        }
    return out


# ---------------------------------------------------------------------------
# Match xlsx rows to DB trades
# ---------------------------------------------------------------------------

def load_db_trades(cur):
    """Return list of dicts for all trades, plus a lookup index by (date, symbol, entry_time)."""
    cur.execute("""
        SELECT id, date, symbol, entry_time, trade_index, legacy_trade_id, net_pnl
        FROM trades
        ORDER BY date, entry_time
    """)
    rows = cur.fetchall()
    by_key = {}
    by_date_symbol = defaultdict(list)
    all_rows = []
    for tid, dt, sym, et, idx, legacy, net in rows:
        rec = {
            "id": tid, "date": dt, "symbol": sym, "entry_time": et,
            "trade_index": idx, "legacy_trade_id": legacy, "net_pnl": float(net or 0),
        }
        all_rows.append(rec)
        by_key[(dt, sym, et)] = rec
        by_date_symbol[(dt, sym)].append(rec)
    return all_rows, by_key, by_date_symbol


def match_xlsx_to_db(execs_xlsx, by_key, by_date_symbol):
    """
    For each xlsx Trade ID, find the DB trade row.
    Primary key: (date, symbol, entry_time) exact match.
    Fallback: (date, symbol) with single match.
    Returns: (matches: dict[xlsx_tid -> db_id], unmatched_xlsx: list, ambiguous: list)
    """
    matches = {}
    unmatched = []
    ambiguous = []

    for tid, ex in execs_xlsx.items():
        key = (ex["date"], ex["symbol"], ex["entry_time"])
        hit = by_key.get(key)
        if hit:
            matches[tid] = hit
            continue
        # Fallback: single trade for that (date, symbol)?
        candidates = by_date_symbol.get((ex["date"], ex["symbol"]), [])
        if len(candidates) == 1:
            matches[tid] = candidates[0]
            matches[tid]["_fallback"] = True
        elif len(candidates) > 1:
            ambiguous.append((tid, ex, candidates))
        else:
            unmatched.append((tid, ex))
    return matches, unmatched, ambiguous


# ---------------------------------------------------------------------------
# Combine Executions + Trades tabs into one record per Trade ID
# ---------------------------------------------------------------------------

def combine_tabs(execs_xlsx, trades_xlsx):
    """
    Flatten the two-tab structure into one dict per TID with all the fields
    the writer needs. Trades tab may be missing for the most-recent un-journaled
    trades; those still get a record with text fields = None.
    """
    out = {}
    for tid, e in execs_xlsx.items():
        t = trades_xlsx.get(tid, {})
        out[tid] = {
            "trade_id":       tid,
            "date":           e["date"],
            "symbol":         e["symbol"],
            "entry_time":     e["entry_time"],
            "side":           e.get("side"),
            "dollar_risk":    e["dollar_risk"],
            "x_failing_goal": e["x_failing_goal"],
            "x_non_playbook": e["x_non_playbook"],
            "x_selection":    e["x_selection"],
            "x_entry":        e["x_entry"],
            "x_sizing":       e["x_sizing"],
            "x_exit":         e["x_exit"],
            "x_emotional":    e["x_emotional"],
            "x_preparation":  e["x_preparation"],
            "setup":          t.get("setup"),
            "sub_setup":      t.get("sub_setup"),
            "trigger":        t.get("trigger"),
            "tags":           t.get("tags"),
            "entry_notes":    t.get("entry_notes"),
            "exit_notes":     t.get("exit_notes"),
            "notes":          t.get("notes"),
            "mistake_notes":  t.get("mistake_notes"),
            "chart":          t.get("chart"),
        }
    return out


# ---------------------------------------------------------------------------
# Writers
# ---------------------------------------------------------------------------

def restamp_legacy_ids(cur, matches):
    """
    Rewrite trades.legacy_trade_id to the xlsx Trade ID for every match.

    Why this dance: the column is UNIQUE, so doing per-row UPDATEs in arbitrary
    order risks transient collisions (e.g. trade A wants ID 1500 currently held
    by trade B; B is about to be set to ID 2000). To dodge that:

      1. Delete the journal rows currently keyed on each match's *old*
         legacy_trade_id (they'll be overwritten by upsert_journal anyway).
      2. NULL out legacy_trade_id on every matched DB trade.
      3. Set each matched trade's legacy_trade_id to the xlsx Trade ID.

    Trades not in `matches` (e.g. the 9 backfilled scratches) keep their IDs.
    """
    matched_db_ids = [m["id"] for m in matches.values()]
    if not matched_db_ids:
        return 0

    # Step 1: delete journal rows for the matched trades' current legacy IDs.
    cur.execute("""
        DELETE FROM trade_journal
        WHERE legacy_trade_id IN (
            SELECT legacy_trade_id FROM trades
            WHERE id = ANY(%s) AND legacy_trade_id IS NOT NULL
        )
    """, (matched_db_ids,))

    # Step 2: clear legacy_trade_id on matched trades so the unique index is free.
    cur.execute("UPDATE trades SET legacy_trade_id = NULL WHERE id = ANY(%s)",
                (matched_db_ids,))

    # Step 3: re-stamp with xlsx IDs.
    n = 0
    for tid, db in matches.items():
        cur.execute(
            "UPDATE trades SET legacy_trade_id = %s WHERE id = %s",
            (tid, db["id"]),
        )
        n += cur.rowcount
    return n


def upsert_journal(cur, tid, m):
    """Upsert one trade_journal row from a merged xlsx record."""
    cur.execute("""
        INSERT INTO trade_journal (
            legacy_trade_id, setup, sub_setup, trigger, tags,
            entry_notes, exit_notes, notes, mistake_notes, chart_url,
            dollar_risk,
            x_failing_goal, x_non_playbook, x_selection, x_entry,
            x_sizing, x_exit, x_emotional, x_preparation
        ) VALUES (
            %s, %s, %s, %s, %s,
            %s, %s, %s, %s, %s,
            %s,
            %s, %s, %s, %s,
            %s, %s, %s, %s
        )
        ON CONFLICT (legacy_trade_id) DO UPDATE SET
            setup          = EXCLUDED.setup,
            sub_setup      = EXCLUDED.sub_setup,
            trigger        = EXCLUDED.trigger,
            tags           = EXCLUDED.tags,
            entry_notes    = EXCLUDED.entry_notes,
            exit_notes     = EXCLUDED.exit_notes,
            notes          = EXCLUDED.notes,
            mistake_notes  = EXCLUDED.mistake_notes,
            chart_url      = EXCLUDED.chart_url,
            dollar_risk    = EXCLUDED.dollar_risk,
            x_failing_goal = EXCLUDED.x_failing_goal,
            x_non_playbook = EXCLUDED.x_non_playbook,
            x_selection    = EXCLUDED.x_selection,
            x_entry        = EXCLUDED.x_entry,
            x_sizing       = EXCLUDED.x_sizing,
            x_exit         = EXCLUDED.x_exit,
            x_emotional    = EXCLUDED.x_emotional,
            x_preparation  = EXCLUDED.x_preparation,
            updated_at     = CURRENT_TIMESTAMP
    """, (
        tid, m["setup"], m["sub_setup"], m["trigger"], m["tags"],
        m["entry_notes"], m["exit_notes"], m["notes"], m["mistake_notes"], m["chart"],
        m["dollar_risk"],
        m["x_failing_goal"], m["x_non_playbook"], m["x_selection"], m["x_entry"],
        m["x_sizing"], m["x_exit"], m["x_emotional"], m["x_preparation"],
    ))


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    args = sys.argv[1:]
    commit = "--commit" in args
    verbose = "--verbose" in args

    if not XLSX_PATH.exists():
        print(f"ERROR: {XLSX_PATH} not found", file=sys.stderr)
        sys.exit(1)

    print(f"Reading {XLSX_PATH}")
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    execs_xlsx = read_executions(wb["Executions"])
    trades_xlsx = read_trades(wb["Trades"])
    print(f"  Executions: {len(execs_xlsx)} trades")
    print(f"  Trades:     {len(trades_xlsx)} trades")

    only_in_exec = set(execs_xlsx) - set(trades_xlsx)
    if only_in_exec:
        print(f"  {len(only_in_exec)} trades in Executions but not Trades (no notes yet): "
              f"{sorted(only_in_exec)[:5]}{'...' if len(only_in_exec) > 5 else ''}")

    merged = combine_tabs(execs_xlsx, trades_xlsx)
    print(f"  Combined records: {len(merged)} trades")

    conn = get_conn()
    try:
        with conn, conn.cursor() as cur:
            all_db, by_key, by_date_symbol = load_db_trades(cur)
            print(f"\nDB trades: {len(all_db)} rows")

            matches, unmatched, ambiguous = match_xlsx_to_db(merged, by_key, by_date_symbol)
            n_fallback = sum(1 for m in matches.values() if m.get("_fallback"))
            print(f"\nMatch summary:")
            print(f"  Matched (exact key):    {len(matches) - n_fallback}")
            print(f"  Matched (fallback):     {n_fallback}")
            print(f"  Unmatched (xlsx -> DB): {len(unmatched)}")
            print(f"  Ambiguous:              {len(ambiguous)}")

            matched_db_ids = {m["id"] for m in matches.values()}
            db_unmatched = [r for r in all_db if r["id"] not in matched_db_ids]
            print(f"  DB trades with no xlsx row: {len(db_unmatched)}")

            if unmatched:
                print(f"\nUnmatched xlsx rows (first 20):")
                for tid, ex in unmatched[:20]:
                    print(f"  TID={tid} {ex['date']} {ex['symbol']:6s} {ex['entry_time']}")
            if ambiguous:
                print(f"\nAmbiguous xlsx rows (first 10):")
                for tid, ex, cands in ambiguous[:10]:
                    print(f"  TID={tid} {ex['date']} {ex['symbol']} "
                          f"xlsx_time={ex['entry_time']} candidates={[c['entry_time'] for c in cands]}")
            if db_unmatched:
                print(f"\nDB trades not in xlsx (first 20):")
                for r in db_unmatched[:20]:
                    print(f"  DB id={r['id']:4d} {r['date']} {r['symbol']:6s} "
                          f"{r['entry_time']} idx={r['trade_index']} net=${r['net_pnl']:+.2f}")

            if verbose:
                print(f"\nFirst 10 matches:")
                for tid in sorted(matches)[:10]:
                    m = matches[tid]
                    fb = " [fallback]" if m.get("_fallback") else ""
                    print(f"  TID {tid} -> DB id {m['id']}  {m['date']} {m['symbol']} {m['entry_time']}{fb}")

            if not commit:
                print("\nDRY RUN - no changes made. Re-run with --commit to write.")
                return

            print(f"\nWriting...")
            n_stamp = restamp_legacy_ids(cur, matches)
            print(f"  Re-stamped legacy_trade_id on {n_stamp} trades")

            n_journal = 0
            for tid in matches:
                upsert_journal(cur, tid, merged[tid])
                n_journal += 1
            print(f"  Wrote/updated {n_journal} trade_journal rows")

            cur.execute("SELECT COUNT(*) FROM trade_journal")
            print(f"  trade_journal now has {cur.fetchone()[0]} rows")
            cur.execute("SELECT COUNT(*) FROM trades WHERE legacy_trade_id IS NOT NULL")
            print(f"  trades with legacy_trade_id: {cur.fetchone()[0]}")
    finally:
        conn.close()
    print("\nDone.")


if __name__ == "__main__":
    main()
